#!/usr/bin/env python3
"""Add MUIO-only structure metadata to an otoole Excel workbook.

This does not change any OSeMOSYS parameter.  It adds technology groups and
human-readable time definitions required by MUIO's existing template importer.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


GROUPS = (
    ("Power generation", "Electricity generation and conversion technologies."),
    ("Energy networks", "Electricity transmission and distribution technologies."),
    ("Energy supply", "Primary energy extraction, imports, exports, and renewable supply."),
    ("Energy demand", "Final energy demand technologies."),
    ("Land and crops", "Land allocation, crop production, and crop processing technologies."),
    ("Water", "Groundwater, surface-water, and public/agricultural water technologies."),
)

REQUIRED_SHEETS = {
    "TECHNOLOGY",
    "TECHGROUP",
    "FUEL",
    "EMISSION",
    "STORAGE",
    "YEAR",
    "MODE_OF_OPERATION",
    "TIMESLICE",
    "SEASON",
    "DAYTYPE",
    "DAILYTIMEBRACKET",
    "InputActivityRatio",
    "OutputActivityRatio",
    "EmissionActivityRatio",
    "TechnologyFromStorage",
    "TechnologyToStorage",
}


def technology_group(name: str) -> str:
    """Return a deterministic display group without changing model behavior."""
    water_prefixes = ("MINPRC", "DEMAGRGWT", "DEMAGRSUR", "DEMPUBGWT", "DEMPUBSUR")
    if name.startswith("PWRTRN"):
        return "Energy networks"
    if name.startswith("PWR"):
        return "Power generation"
    if name.startswith(water_prefixes):
        return "Water"
    if name.startswith(("LND", "MINLND")):
        return "Land and crops"
    if name.startswith("DEM"):
        return "Energy demand"
    if name.startswith(("MIN", "IMP", "EXP", "RNW")):
        return "Energy supply"
    raise ValueError(f"No TECHGROUP rule for technology {name!r}")


def column_map(worksheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }


def ensure_column(worksheet, name: str) -> int:
    columns = column_map(worksheet)
    if name in columns:
        return columns[name]
    column = worksheet.max_column + 1
    worksheet.cell(1, column, name)
    return column


def replace_techgroup_sheet(workbook) -> None:
    if "TECHGROUP" in workbook.sheetnames:
        del workbook["TECHGROUP"]
    position = workbook.sheetnames.index("TECHNOLOGY") + 1
    worksheet = workbook.create_sheet("TECHGROUP", position)
    worksheet.append(("TECHGROUP", "DESCRIPTION"))
    for row in GROUPS:
        worksheet.append(row)


def annotate_technologies(workbook) -> None:
    worksheet = workbook["TECHNOLOGY"]
    value_column = column_map(worksheet)["VALUE"]
    group_column = ensure_column(worksheet, "TECHGROUP")
    description_column = ensure_column(worksheet, "DESCRIPTION")
    assigned_groups: set[str] = set()

    for row in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row, value_column).value
        if value is None:
            continue
        name = str(value).strip()
        group = technology_group(name)
        assigned_groups.add(group)
        worksheet.cell(row, group_column, group)
        worksheet.cell(row, description_column, f"{name} ({group.lower()}).")

    declared_groups = {name for name, _ in GROUPS}
    if assigned_groups != declared_groups:
        missing = sorted(declared_groups - assigned_groups)
        raise ValueError(f"Declared TECHGROUP values without technologies: {missing}")


def add_descriptions(workbook) -> None:
    descriptions = {
        "TIMESLICE": {
            "S1D1": "Wet season, day.",
            "S1D2": "Wet season, night.",
            "S2D1": "Dry season, day.",
            "S2D2": "Dry season, night.",
        },
        "SEASON": {"1": "Wet season.", "2": "Dry season."},
        "DAYTYPE": {"1": "Representative day type."},
        "DAILYTIMEBRACKET": {"1": "Day.", "2": "Night."},
    }
    for sheet_name, values in descriptions.items():
        worksheet = workbook[sheet_name]
        value_column = column_map(worksheet)["VALUE"]
        description_column = ensure_column(worksheet, "DESCRIPTION")
        seen: set[str] = set()
        for row in range(2, worksheet.max_row + 1):
            raw_value = worksheet.cell(row, value_column).value
            if raw_value is None:
                continue
            key = str(raw_value).removesuffix(".0")
            if key not in values:
                raise ValueError(f"Unexpected {sheet_name} value {raw_value!r}")
            worksheet.cell(row, description_column, values[key])
            seen.add(key)
        if seen != set(values):
            raise ValueError(
                f"{sheet_name} values differ from the expected Philippines definitions"
            )


def ensure_discount_rate(workbook, fallback: float = 0.05) -> bool:
    """Make the agreed fallback explicit when CLEWs Global supplies no row."""
    worksheet = workbook["DiscountRate"]
    has_data = any(
        cell.value is not None
        for row in worksheet.iter_rows(min_row=2)
        for cell in row
    )
    if has_data:
        return False
    columns = column_map(worksheet)
    worksheet.cell(2, columns["REGION"], "GLOBAL")
    worksheet.cell(2, columns["VALUE"], fallback)
    return True


def remove_empty_optional_sheets(workbook) -> list[str]:
    """Omit empty otoole parameter sheets that ImportTemplate treats as populated."""
    removed: list[str] = []
    for sheet_name in list(workbook.sheetnames):
        if sheet_name in REQUIRED_SHEETS:
            continue
        worksheet = workbook[sheet_name]
        has_data = any(
            cell.value is not None
            for row in worksheet.iter_rows(min_row=2)
            for cell in row
        )
        if not has_data:
            del workbook[sheet_name]
            removed.append(sheet_name)
    return removed


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path, help="otoole-generated .xlsx workbook")
    parser.add_argument("output", type=Path, help="MUIO-ready .xlsx workbook")
    args = parser.parse_args()

    workbook = load_workbook(args.input)
    replace_techgroup_sheet(workbook)
    annotate_technologies(workbook)
    add_descriptions(workbook)
    used_discount_rate_fallback = ensure_discount_rate(workbook)
    removed = remove_empty_optional_sheets(workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Prepared MUIO workbook: {args.output}")
    if used_discount_rate_fallback:
        print("Applied DiscountRate fallback: 0.05 (5%)")
    print(f"Omitted {len(removed)} empty optional parameter sheets: {', '.join(removed)}")


if __name__ == "__main__":
    main()
