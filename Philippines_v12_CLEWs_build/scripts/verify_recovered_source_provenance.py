#!/usr/bin/env python3
"""Read-only verification of the recovered Philippines v12 source evidence."""

from __future__ import annotations

import csv
import hashlib
import math
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
EVIDENCE = ROOT / "data_sources" / "evidence" / "land_agriculture_water"

EXPECTED_HASHES = {
    "gaez_input_tables/GAEZ_cwd_High_Input.csv": "900596e1d956adc18e2c77cbc3a07e50c7363f72897f350b88c6a2a2e7db3e9a",
    "gaez_input_tables/GAEZ_cwd_Low_Input.csv": "6817b8d2561fac0bb7bfc38ba8076a50fd4d9e449815b78dfed8a9bb827afcd1",
    "gaez_input_tables/GAEZ_evt_High_Input.csv": "5e1df43709dfc16d94e690c0bacd76e8829d8e86cf747322a52467c7296e7cbc",
    "gaez_input_tables/GAEZ_evt_Low_Input.csv": "c931c0d6c6dffff7936f3835f872560576ca6f739c2c99cea713bac92220a874",
    "gaez_input_tables/GAEZ_yld_High_Input.csv": "47121388a94a61cd7158e9f2ad9474d3480d93b2aa292e154868760bcf84a531",
    "gaez_input_tables/GAEZ_yld_Low_Input.csv": "051c31b303930775bbe12229fd59b9257eed5d36b5a01c3249d4655e498a8beb",
    "gaez_input_tables/GAEZ_PHL_RASTER_CACHE_MANIFEST.csv": "d9325d79f75dd4dcbd26791aea133de811fbc217b783990a18d24b4959e87f26",
    "gaez_base_rasters/precipitation prc.tif": "c72b8cec38ea31755b9d6f02e518571f41f7d39ae2f5913ee64a14077079d83b",
    "gaez_base_rasters/LCType_ncb.tif": "17b77aa5fd7b56a4570119026034c0f3137b9e906c741c684556f762c34a30db",
    "gaez_base_rasters/GAEZ_BUNDLED_BASE_RASTERS.csv": "48745a54c23e22cde7ae82a69a77abbaa8bc6ed48f9aca0cfc7cfff54dd56486",
    "faostat/FAOSTAT_2020.csv": "082da2e1c708ea6f9837b021143eb193d9f240b0700be367d80ab91cfeb259b9",
    "faostat/FAOSTAT_production_2020.csv": "51fe2f975b0d57ddcca0e4da860b12c89b7aa9a9daa41b3fe13df77fd746a14f",
    "faostat/FAOSTAT_PHL_2020_SELECTION.csv": "f1aff540fc2a3187548a9cc6b79b602baa7a6031b73e44c633974b56a31920bf",
    "ssp2/iamc_db_POP_Countries.xlsx": "e013c76eacd95a2e0f6275446e9d1e71bdfb9556eab1c6619a5b0d7914ecab35",
    "ssp2/SSP2_PHL_SELECTED_ROW.csv": "a380ec78c557c93d491decfda48f41e2ac928c8776cd196c5f5ddb4d13f0d07c",
    "ssp2/SSP2_PHL_ANNUAL_INDEX_2020_2053.csv": "a9c7297c06d33e07c512a38fd921eb02a7517a18a27e2cfef16f9fb9f8fd343c",
    "workflow/WORKFLOW_SESSION_EVENTS.csv": "b553b2939790a855c010bb5491e8052d7f653c4c851f5d102eaa5ed397f49250",
}

SESSION = Path(
    "/Users/sato/.codex/sessions/2026/07/25/"
    "rollout-2026-07-25T09-57-55-019f9991-7a2d-72a0-8fa1-6820178dd177.jsonl"
)
SESSION_HASH = "b472410035d9f1eb61e3142a3f5b9311c8ec8a872d2b6d83728fd39b18555f65"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def close(actual: str, expected: float) -> bool:
    return math.isclose(float(actual), expected, rel_tol=0, abs_tol=1e-10)


def main() -> int:
    failures: list[str] = []
    for relative, expected in EXPECTED_HASHES.items():
        path = EVIDENCE / relative
        if not path.is_file():
            failures.append(f"missing {relative}")
        elif sha256(path) != expected:
            failures.append(f"hash mismatch {relative}")

    manifest = rows(EVIDENCE / "gaez_input_tables/GAEZ_PHL_RASTER_CACHE_MANIFEST.csv")
    levels = {level: sum(r["input_level"] == level for r in manifest) for level in ("High", "Low")}
    if len(manifest) != 90 or levels != {"High": 48, "Low": 42}:
        failures.append(f"GAEZ manifest count mismatch: total={len(manifest)}, levels={levels}")
    for field in ("cache_filename", "download_url"):
        if len({r[field] for r in manifest}) != 90:
            failures.append(f"GAEZ manifest {field} is not unique")
    low_rubber = [r for r in manifest if r["input_level"] == "Low" and r["model_crop_code"] == "RUB"]
    if low_rubber:
        failures.append("unexpected low-input RUB rows")

    selected = rows(EVIDENCE / "faostat/FAOSTAT_PHL_2020_SELECTION.csv")
    expected_fao = [
        ("0113", 7093, 4718896.0, 7251, 19294855.51),
        ("01460", 7052, 3651289.0, 7209, 14490922.69),
        ("0112", 7065, 2553781.0, 7222, 8118545.88),
        ("01290.90", 7081, 639800.0, 7239, 5498265.59),
        ("01802", 7100, 399086.0, 7258, 24398941.25),
        ("01319", 7080, 375101.0, 7238, 3215377.53),
        ("01313", 7088, 263580.0, 7246, 3100838.72),
        ("01950.01", 7067, 230723.0, 7225, 422407.1),
        ("01520.01", 7046, 219348.0, 7203, 2607758.63),
        ("01316", 7066, 195135.0, 7223, 753103.14),
    ]
    if len(selected) != len(expected_fao):
        failures.append(f"FAOSTAT selected-row count mismatch: {len(selected)}")
    else:
        for row, expected in zip(selected, expected_fao):
            code, area_line, area, production_line, production = expected
            if not (
                row["Item Code (CPC)"] == code
                and int(row["_source_csv_line_harvested_area"]) == area_line
                and close(row["Value_harvested_area"], area)
                and int(row["_source_csv_line_production"]) == production_line
                and close(row["Value_production"], production)
            ):
                failures.append(f"FAOSTAT selected-row mismatch for {code}")

    selected_ssp = rows(EVIDENCE / "ssp2/SSP2_PHL_SELECTED_ROW.csv")
    if len(selected_ssp) != 1:
        failures.append("SSP2 selected-row CSV does not contain exactly one row")
    else:
        row = selected_ssp[0]
        expected_identity = {
            "sheet": "data", "excel_row": "1373", "Model": "IIASA-WiC POP",
            "Scenario": "SSP2", "Region": "PHL", "Variable": "Population", "Unit": "million",
        }
        if any(row[key] != value for key, value in expected_identity.items()):
            failures.append("SSP2 selected-row identity mismatch")

    annual = rows(EVIDENCE / "ssp2/SSP2_PHL_ANNUAL_INDEX_2020_2053.csv")
    annual_by_year = {int(r["year"]): r for r in annual}
    if len(annual) != 34 or set(annual_by_year) != set(range(2020, 2054)):
        failures.append("SSP2 annual index year coverage mismatch")
    else:
        for year in (2050, 2051, 2052, 2053):
            if not close(annual_by_year[year]["population_million"], 149.04005805640713):
                failures.append(f"SSP2 trailing-fill mismatch for {year}")
        if not close(annual_by_year[2020]["population_index_2020"], 1.0):
            failures.append("SSP2 normalization mismatch for 2020")

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(
            EVIDENCE / "ssp2/iamc_db_POP_Countries.xlsx", read_only=True, data_only=True
        )
        if workbook.sheetnames != ["data", "Recommended Citation"]:
            failures.append(f"SSP2 workbook sheets mismatch: {workbook.sheetnames}")
        values = list(workbook["data"].iter_rows(min_row=1373, max_row=1373, values_only=True))[0]
        if tuple(values[:5]) != ("IIASA-WiC POP", "SSP2", "PHL", "Population", "million"):
            failures.append("SSP2 workbook data!A1373:E1373 mismatch")
        workbook.close()
    except ImportError:
        print("WARN: openpyxl unavailable; workbook bytes and extracted selected row were verified")

    if SESSION.is_file():
        if sha256(SESSION) != SESSION_HASH:
            failures.append("raw build-session hash mismatch")
    else:
        print(f"WARN: raw build session not available at {SESSION}")

    if failures:
        for failure in failures:
            print(f"FAIL: {failure}")
        return 1
    print("PASS: recovered Philippines v12 source evidence is internally consistent")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
