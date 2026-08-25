#!/usr/bin/env python3
"""Create the review workbook from the authoritative Philippines v18 CSVs."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PACKAGE = Path(__file__).resolve().parents[1]
LEDGER = PACKAGE / "data_sources"
OUTPUT = LEDGER / "PHILIPPINES_V18_CANONICAL_SCHEMA_LEDGER.xlsx"
TABLES = [
    "SOURCES.csv", "CALCULATIONS.csv", "ASSUMPTIONS.csv",
    "MODEL_MAP.csv", "GAPS.csv", "CHANGES.csv",
]


def row_count(path: Path) -> int:
    with path.open(newline="", encoding="utf-8-sig") as stream:
        return sum(1 for _ in csv.reader(stream)) - 1


def main() -> None:
    workbook = Workbook(write_only=True)
    index = workbook.create_sheet("INDEX")
    index.append(["Philippines v18 canonical cumulative schema ledger"])
    index.append(["Authority", "The six CSV files beside this workbook are authoritative; this workbook is a review copy."])
    index.append(["Scope", "Inherited base + v13 calibration + v14 stock turnover + v15 national water + complete v16 repairs + v17 national land-cover accounting + v18 energy inputs and deployment envelopes"])
    index.append(["Self-containment", "No earlier installed case or earlier-version ledger is required."])
    index.append([])
    index.append(["Sheet", "Rows"])
    for filename in TABLES:
        index.append([filename.removesuffix(".csv"), row_count(LEDGER / filename)])
    index.column_dimensions["A"].width = 28
    index.column_dimensions["B"].width = 95
    index.freeze_panes = "A7"

    for filename in TABLES:
        worksheet = workbook.create_sheet(filename.removesuffix(".csv")[:31])
        with (LEDGER / filename).open(newline="", encoding="utf-8-sig") as stream:
            reader = csv.reader(stream)
            header = next(reader)
            styled_header = []
            widths = [len(value) for value in header]
            for value in header:
                cell = WriteOnlyCell(worksheet, value=value)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                styled_header.append(cell)
            worksheet.append(styled_header)
            count = 1
            for row in reader:
                worksheet.append(row)
                count += 1
                for index_value, value in enumerate(row):
                    widths[index_value] = min(max(widths[index_value], len(str(value))), 60)
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(header))}{count}"
            for index_value, width in enumerate(widths, start=1):
                worksheet.column_dimensions[get_column_letter(index_value)].width = max(12, min(width + 2, 60))

    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
