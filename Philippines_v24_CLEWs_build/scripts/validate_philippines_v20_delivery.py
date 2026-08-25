#!/usr/bin/env python3
"""Validate the self-contained Philippines v20 ledger and result-free delivery."""

from __future__ import annotations

import csv
import hashlib
import json
import zipfile
from pathlib import Path

from openpyxl import load_workbook


PACKAGE = Path(__file__).resolve().parents[1]
REPO = PACKAGE.parent
LEDGER = PACKAGE / "data_sources"
CASE = REPO / "case" / "Philippines_v20"
OUTPUT = PACKAGE / "diagnostics" / "power_history_schema_ledger_validation.json"
TABLES = ["SOURCES.csv", "CALCULATIONS.csv", "ASSUMPTIONS.csv", "MODEL_MAP.csv", "GAPS.csv", "CHANGES.csv"]
KEYS = {"SOURCES.csv": "source_id", "CALCULATIONS.csv": "calculation_id", "ASSUMPTIONS.csv": "assumption_id", "MODEL_MAP.csv": "map_id", "GAPS.csv": "item", "CHANGES.csv": "change_id"}
SOURCE_FILES = ["R.json", "RE.json", "RS.json", "RT.json", "RTSM.json", "RYC.json", "RYCTs.json", "RYCn.json", "RYDtb.json", "RYE.json", "RYS.json", "RYSeDt.json", "RYT.json", "RYTC.json", "RYTCM.json", "RYTCn.json", "RYTEM.json", "RYTM.json", "RYTTs.json", "RYTs.json", "genData.json"]
V20_SOURCE_IDS = {
    "SRC_PHL_DOE_POWER_SUMMARY_2024", "SRC_PHL_DOE_POWER_CAPACITY_2020_2024",
    "SRC_PHL_DOE_EXISTING_PLANTS_2020", "SRC_PHL_NATGAS_MASTER_PLAN",
    "SRC_PHL_PEP_2023_2050_VOL2", "SRC_PHL_V20_POWER_BUILD",
    "SRC_PHL_V20_POWER_VALIDATION", "SRC_PHL_V20_MODEL_ARCHIVE",
}


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_table(name: str) -> list[dict[str, str]]:
    with (LEDGER / name).open(newline="", encoding="utf-8-sig") as stream:
        rows = list(csv.DictReader(stream))
    values = [row[KEYS[name]] for row in rows]
    assert len(values) == len(set(values)), name
    return rows


def refs(value: str) -> set[str]:
    return set(value.replace(";", " ").split())


def row(table: dict, parameter: str, technology: str, mode: int | None = None) -> dict:
    matches = [item for item in table[parameter]["SC_0"] if item.get("TechId") == technology and (mode is None or item.get("MoId") == mode)]
    assert len(matches) == 1, (parameter, technology, mode)
    return matches[0]


def main() -> None:
    tables = {name: read_table(name) for name in TABLES}
    source_ids = {item["source_id"] for item in tables["SOURCES.csv"]}
    calculation_ids = {item["calculation_id"] for item in tables["CALCULATIONS.csv"]}
    assumption_ids = {item["assumption_id"] for item in tables["ASSUMPTIONS.csv"]}
    evidence_ids = source_ids | calculation_ids | assumption_ids

    for item in tables["SOURCES.csv"]:
        if item["local_file"]:
            path = LEDGER / item["local_file"]
            assert path.is_file(), path
            if item["sha256"] and item["source_id"] in V20_SOURCE_IDS:
                assert digest(path) == item["sha256"], item["source_id"]
    for item in tables["CALCULATIONS.csv"]:
        assert refs(item["source_ids"]) <= source_ids
        assert refs(item["assumption_ids"]) <= assumption_ids
        assert refs(item["input_calculation_ids"]) <= calculation_ids
    for item in tables["ASSUMPTIONS.csv"]:
        assert refs(item["evidence_source_ids"]) <= source_ids
    for item in tables["MODEL_MAP.csv"]:
        assert refs(item["evidence_ids"]) <= evidence_ids, item["map_id"]

    ryt, rytm = json.loads((CASE / "RYT.json").read_text()), json.loads((CASE / "RYTM.json").read_text())
    expected_af = {"TEC_pyjfk": 0.9361653523880883, "TEC_ze7d4": 0.9518030412744388, "TEC_2hnym": 0.7207666525043667, "TEC_gthhk": 0.6379079123826553}
    for technology, expected in expected_af.items():
        actual = row(ryt, "AF", technology)
        assert all(actual[str(year)] == expected for year in range(2020, 2054))
    assert row(ryt, "FC", "TEC_ze7d4")["2020"] == 320.93815402543083
    assert row(ryt, "FC", "TEC_ze7d4")["2021"] == 278.39368962952454
    assert row(rytm, "VC", "TEC_ze7d4", 1)["2020"] == -13.16562277957841
    assert row(rytm, "VC", "TEC_ze7d4", 1)["2021"] == -13.216931130239278
    assert row(rytm, "VC", "TEC_6rwx6", 1)["2020"] == 6.662111756970275
    assert row(rytm, "VC", "TEC_6rwx6", 1)["2021"] == 6.68807478018528

    static = json.loads((LEDGER / "snapshots" / "power_calibration_v20_static_validation.json").read_text())
    validation = json.loads((LEDGER / "snapshots" / "power_calibration_v20_validation.json").read_text())
    promotion = json.loads((LEDGER / "snapshots" / "power_calibration_v20_promotion_identity.json").read_text())
    assert static["status"] == "passed"
    assert validation["status"] == "accepted"
    assert validation["observation_classification"]["forcing"] == "none"
    assert validation["sensitivity_runs"] == 0
    assert promotion["status"] == "passed" and promotion["optimizer_runs"] == 0

    manifest_path = LEDGER / "V20_MODEL_ARCHIVE_MANIFEST.csv"
    with manifest_path.open(newline="", encoding="utf-8-sig") as stream:
        manifest = next(csv.DictReader(stream))
    archive = PACKAGE / "muio" / manifest["archive"]
    assert digest(archive) == manifest["sha256"]
    assert archive.stat().st_size == int(manifest["size_bytes"])
    with zipfile.ZipFile(archive) as bundle:
        members = [info.filename for info in bundle.infolist() if not info.is_dir()]
        assert len(members) == int(manifest["member_count"])
        assert all(name.startswith("Philippines_v20/") for name in members)
        assert not any("/res/" in name or Path(name).name in {"data.txt", "data_processed.txt", "lp.lp", "results.txt"} for name in members)

    workbook_path = LEDGER / "PHILIPPINES_V20_CANONICAL_SCHEMA_LEDGER.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assert set(["INDEX"] + [name.removesuffix(".csv") for name in TABLES]) <= set(workbook.sheetnames)
    for name in TABLES:
        workbook_rows = sum(1 for _ in workbook[name.removesuffix(".csv")].iter_rows(values_only=True)) - 1
        assert workbook_rows == len(tables[name]), name
    workbook.close()

    current_manifest = LEDGER / "evidence" / "CURRENT_MODEL_ARCHIVE_MANIFEST.csv"
    assert archive.name in current_manifest.read_text(encoding="utf-8-sig")
    assert all((CASE / name).is_file() for name in SOURCE_FILES)
    report = {
        "schema": "philippines-v20-power-history-delivery-validation-v1",
        "status": "passed",
        "ledger_rows": {name: len(rows) for name, rows in tables.items()},
        "v20_records": {"sources": 8, "calculations": 8, "assumptions": 4, "maps": 11, "gaps": 5, "changes": 1},
        "source_parameter_checks": "passed",
        "cross_references_and_retained_hashes": "passed",
        "review_workbook": str(workbook_path),
        "archive": {"path": str(archive), "sha256": digest(archive), "members": len(members), "result_files_included": False},
        "promotion_identity": promotion,
        "optimizer_runs": {"rejected_diagnostic": 1, "accepted_corrective": 1, "post_promotion": 0, "sensitivities": 0},
    }
    OUTPUT.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
