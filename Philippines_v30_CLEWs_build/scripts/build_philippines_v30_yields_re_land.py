#!/usr/bin/env python3
"""Build the standalone Philippines_v30 source package from canonical v29.

V30 removes the unsupported crop input-level choice, initializes yields only in
2020 and evolves them with FAO's independent Business As Usual crop projection,
and makes population, solar-PV, and onshore-wind land compete inside the
existing physical built-up land account.

This builder never invokes GLPK or CBC.
"""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from build_philippines_v29_crop_land import CLUSTERS, OLD_OPTIONS


ROOT = Path(__file__).resolve().parents[1]
STORAGE = ROOT / "WebAPP" / "DataStorage"
SOURCE = STORAGE / "Philippines_v29"
TARGET = STORAGE / "Philippines_v30"
BASE = "SC_0"
YEARS = [str(year) for year in range(2020, 2054)]
MODES = range(1, 31)

BUILT = "TEC_bn8d7"
BUILT_LAND = "COM_zxzvt"
BUILT_SITE = "COM_phl_built_site_v30"
PV = "TEC_1k064"
WIND = "TEC_1wdli"
CONSTRAINT = "CO_phl_built_v30"
PV_FOOTPRINT = 0.012
WIND_FOOTPRINT = 0.010
PV_2020 = 1.01207
WIND_2020 = 0.4269
OBSERVED_BUILT_2020 = 10.2649
POPULATION_BUILT_2020 = OBSERVED_BUILT_2020 - PV_FOOTPRINT * PV_2020 - WIND_FOOTPRINT * WIND_2020
INACTIVE_SENTINEL = 0.000001

ACTIVE_MODES = {1, 3, 4, 5, 6, 8, 10, 11, 16, 19, 22, 24}
INACTIVE_MODES = set(range(1, 25)) - ACTIVE_MODES
CROP_COMMODITIES = {
    "vegetables": "COM_mmit0", "coconut": "COM_vwhhn", "sugarcane": "COM_ec7z7",
    "maize": "COM_v6c8s", "rice_rain": "COM_zrfky", "rice_irrig": "COM_zrfky",
    "other": "COM_np43p",
}
FOFA_SOURCE_URL = "https://www.fao.org/fileadmin/user_upload/global-perspective/csv/FOFA2050CountryData_Crop-production.csv.zip"
FOFA_SOURCE_SHA256 = "8b1f1868155b4446dc30193faafa876962544ebd2f44cff50ade987018942084"
FOFA = {
    ("bananas", "rainfed"): {2012: 20.022, 2030: 25.384, 2035: 26.351, 2040: 27.148, 2050: 28.273},
    ("cassava", "rainfed"): {2012: 10.364, 2030: 13.838, 2035: 14.49, 2040: 15.033, 2050: 15.804},
    ("coconut", "rainfed"): {2012: 4.348, 2030: 5.046, 2035: 5.198, 2040: 5.338, 2050: 5.57},
    ("maize", "rainfed"): {2012: 2.824, 2030: 3.924, 2035: 4.218, 2040: 4.501, 2050: 5.012},
    ("rubber", "irrigated"): {2012: 0.697, 2030: 0.786, 2035: 0.806, 2040: 0.825, 2050: 0.855},
    ("other_fruits", "rainfed"): {2012: 9.213, 2030: 9.831, 2035: 9.997, 2040: 10.172, 2050: 10.514},
    ("vegetables", "irrigated"): {2012: 10.793, 2030: 12.949, 2035: 13.485, 2040: 13.999, 2050: 14.902},
    ("vegetables", "rainfed"): {2012: 8.541, 2030: 10.248, 2035: 10.672, 2040: 11.079, 2050: 11.794},
    ("rice", "irrigated"): {2012: 4.419, 2030: 5.399, 2035: 5.628, 2040: 5.838, 2050: 6.184},
    ("rice", "rainfed"): {2012: 3.192, 2030: 3.874, 2035: 4.031, 2040: 4.173, 2050: 4.403},
    ("sugarcane", "irrigated"): {2012: 73.308, 2030: 84.919, 2035: 87.157, 2040: 89.145, 2050: 92.095},
    ("sugarcane", "rainfed"): {2012: 70.071, 2030: 80.455, 2035: 82.368, 2040: 84.033, 2050: 86.366},
}
OTHER_PRODUCTION_WEIGHTS = {
    "other_fruits": 3215377.53 + 753103.14,
    "bananas": 3100838.72,
    "rubber": 422407.10,
    "cassava": 2607758.63,
}
LEDGERS = ["SOURCES.csv", "ASSUMPTIONS.csv", "CALCULATIONS.csv", "MODEL_MAP.csv", "CHANGES.csv", "GAPS.csv"]
OTHER_LAND_TECH = "TEC_zty92"


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def append_ledger(filename: str, rows: list[dict[str, str]], key: str) -> None:
    path = TARGET / "data_sources" / filename
    fields, existing = read_csv(path)
    known = {row[key] for row in existing}
    for row in rows:
        if set(row) != set(fields):
            raise AssertionError(f"{filename} schema mismatch: {set(row) ^ set(fields)}")
        if row[key] in known:
            raise AssertionError(f"duplicate ledger ID {row[key]}")
        existing.append(row)
        known.add(row[key])
    write_csv(path, fields, existing)


def row(values: list[str], fields: list[str]) -> dict[str, str]:
    if len(values) != len(fields):
        raise AssertionError((len(values), len(fields), fields, values))
    return dict(zip(fields, values))


def find_row(rows: list[dict[str, Any]], **keys: Any) -> dict[str, Any]:
    found = [item for item in rows if all(item.get(key) == value for key, value in keys.items())]
    if len(found) != 1:
        raise AssertionError(f"expected one row for {keys}, found {len(found)}")
    return found[0]


def values(value: Any) -> dict[str, Any]:
    return {year: value for year in YEARS}


def projected_value(points: dict[int, float], year: int) -> float:
    """Linearly interpolate FAO milestones and extend the 2040-2050 slope."""
    if year > 2050:
        return points[2050] + (year - 2050) * (points[2050] - points[2040]) / 10.0
    knots = sorted(points)
    for left, right in zip(knots, knots[1:]):
        if left <= year <= right:
            fraction = (year - left) / (right - left)
            return points[left] + fraction * (points[right] - points[left])
    raise AssertionError(year)


def fofa_factor(crop: str, water: str, year: int) -> float:
    if crop.startswith("rice_"):
        key = ("rice", water)
        return projected_value(FOFA[key], year) / projected_value(FOFA[key], 2020)
    if crop in {"maize", "coconut"}:
        key = (crop, "rainfed")
        return projected_value(FOFA[key], year) / projected_value(FOFA[key], 2020)
    if crop in {"vegetables", "sugarcane"}:
        key = (crop, water)
        return projected_value(FOFA[key], year) / projected_value(FOFA[key], 2020)
    if crop == "other":
        numerator = 0.0
        denominator = sum(OTHER_PRODUCTION_WEIGHTS.values())
        for component, weight in OTHER_PRODUCTION_WEIGHTS.items():
            regime = "irrigated" if component == "rubber" else "rainfed"
            points = FOFA[(component, regime)]
            numerator += weight * projected_value(points, year) / projected_value(points, 2020)
        return numerator / denominator
    raise AssertionError((crop, water))


def copy_parent() -> None:
    if not SOURCE.is_dir():
        raise SystemExit(f"missing canonical parent {SOURCE}")
    if TARGET.exists():
        raise SystemExit(f"refusing to overwrite {TARGET}")
    TARGET.mkdir()
    for item in SOURCE.iterdir():
        if item.name in {"res", "view"}:
            continue
        destination = TARGET / item.name
        shutil.copytree(item, destination, symlinks=True) if item.is_dir() else shutil.copy2(item, destination)
    (TARGET / "res").mkdir()
    (TARGET / "view").mkdir()
    write_json(TARGET / "view" / "resData.json", {"osy-cases": []})


def population_built_path() -> dict[str, float]:
    rytm = read_json(SOURCE / "RYTM.json")
    inherited = find_row(rytm["TAMLL"][BASE], TechId=BUILT, MoId=1)
    scale = POPULATION_BUILT_2020 / OBSERVED_BUILT_2020
    return {year: float(inherited[year]) * scale for year in YEARS}


def install_general_data() -> None:
    gen = read_json(TARGET / "genData.json")
    gen["osy-casename"] = "Philippines_v30"
    gen["osy-desc"] = (
        "Philippines v30 crop-yield and renewable-land repair. Only one evidenced management system per crop and "
        "water regime is operable; yields are initialized in 2020 and subsequently follow the FAO 2050 Business As Usual projection. "
        "Population, solar PV and onshore wind consume a shared built-site service produced through the existing "
        "physical built-up land chain, with an exact endogenous annual land identity.\n\n" + gen["osy-desc"]
    )
    gen["osy-comm"].append({
        "CommId": BUILT_SITE, "Comm": "PHL_BUILT_SITE",
        "Desc": "Built-site service shared by population settlement, solar PV and onshore wind; supplied only through the existing LBLTTOT spatial land modes.",
        "UnitId": "10<sup>3</sup>km<sup>2</sup>",
    })
    techs = {item["TechId"]: item for item in gen["osy-tech"]}
    for cluster in CLUSTERS:
        techs[cluster]["OAR"] = sorted(set(techs[cluster]["OAR"] + [BUILT_SITE]))
        techs[cluster]["Desc"] += " V30 mode 26 converts physical built-up land into the shared population/PV/wind site service."
    techs[PV]["ITCR"] = sorted(set(techs[PV]["ITCR"] + [BUILT_SITE]))
    techs[WIND]["ITCR"] = sorted(set(techs[WIND]["ITCR"] + [BUILT_SITE]))
    gen["osy-constraints"].append({
        "ConId": CONSTRAINT, "Con": "PHL_DYNAMIC_BUILT_LAND",
        "Desc": "Exact annual built-land identity: built activity equals population settlement land plus PV and onshore-wind installed-capacity footprints.",
        "Tag": 1, "CM": [BUILT, PV, WIND],
    })
    write_json(TARGET / "genData.json", gen)


def install_built_land(population: dict[str, float]) -> None:
    gen = read_json(TARGET / "genData.json")
    scenarios = [item["ScenarioId"] for item in gen["osy-scenarios"]]

    rytm = read_json(TARGET / "RYTM.json")
    for parameter, value in (("TAMLL", 0), ("TAMUL", 99999)):
        find_row(rytm[parameter][BASE], TechId=BUILT, MoId=1).update(values(value))
    # Clear the obsolete v28 policy-only idle-cropland cap so policies inherit BASE.
    for scenario in scenarios:
        if scenario == BASE:
            continue
        for parameter in ("TAMLL", "TAMUL"):
            find_row(rytm[parameter][scenario], TechId=OTHER_LAND_TECH, MoId=2).update(values(None))
    write_json(TARGET / "RYTM.json", rytm)

    rytcm = read_json(TARGET / "RYTCM.json")
    for scenario in scenarios:
        for cluster in CLUSTERS:
            for mode in MODES:
                cell = 1.0 if scenario == BASE and mode == 26 else (0.0 if scenario == BASE else None)
                rytcm["OAR"][scenario].append({"TechId": cluster, "CommId": BUILT_SITE, "MoId": mode, **values(cell)})
    write_json(TARGET / "RYTCM.json", rytcm)

    rytc = read_json(TARGET / "RYTC.json")
    for scenario in scenarios:
        for tech, coefficient in ((PV, PV_FOOTPRINT), (WIND, WIND_FOOTPRINT)):
            rytc["ITCR"][scenario].append({
                "TechId": tech, "CommId": BUILT_SITE,
                **values(coefficient if scenario == BASE else None),
            })
    write_json(TARGET / "RYTC.json", rytc)

    ryc = read_json(TARGET / "RYC.json")
    for parameter in ryc:
        for scenario in scenarios:
            value = population if parameter == "AAD" and scenario == BASE else values(0 if scenario == BASE else None)
            ryc[parameter][scenario].append({"CommId": BUILT_SITE, **value})
    write_json(TARGET / "RYC.json", ryc)

    rycts = read_json(TARGET / "RYCTs.json")
    for parameter in rycts:
        for scenario in scenarios:
            for timeslice in gen["osy-ts"]:
                rycts[parameter][scenario].append({
                    "CommId": BUILT_SITE, "TsId": timeslice["TsId"],
                    **values(0 if scenario == BASE else None),
                })
    write_json(TARGET / "RYCTs.json", rycts)

    rytcn = read_json(TARGET / "RYTCn.json")
    for parameter in ("CAM", "CCM", "CNCM"):
        for scenario in scenarios:
            for tech in (BUILT, PV, WIND):
                coefficient = 0.0
                if parameter == "CAM" and tech == BUILT:
                    coefficient = 1.0
                if parameter == "CCM" and tech == PV:
                    coefficient = -PV_FOOTPRINT
                if parameter == "CCM" and tech == WIND:
                    coefficient = -WIND_FOOTPRINT
                rytcn[parameter][scenario].append({
                    "TechId": tech, "ConId": CONSTRAINT,
                    **values(coefficient if scenario == BASE else None),
                })
    write_json(TARGET / "RYTCn.json", rytcn)

    rycn = read_json(TARGET / "RYCn.json")
    for parameter in rycn:
        for scenario in scenarios:
            value = population if parameter == "UCC" and scenario == BASE else values(0 if scenario == BASE else None)
            rycn[parameter][scenario].append({"ConId": CONSTRAINT, **value})
    write_json(TARGET / "RYCn.json", rycn)


def install_yield_repair() -> list[dict[str, Any]]:
    rytm = read_json(TARGET / "RYTM.json")
    rytcm = read_json(TARGET / "RYTCM.json")
    mapping = {mode: (crop, level, water) for mode, _ti, _tn, _ci, _cn, crop, level, water in OLD_OPTIONS}
    evidence: list[dict[str, Any]] = []

    for cluster_id, cluster_name in CLUSTERS.items():
        for mode in range(1, 25):
            crop, level, water = mapping[mode]
            upper = find_row(rytm["TAMUL"][BASE], TechId=cluster_id, MoId=mode)
            upper.update(values(99999 if mode in ACTIVE_MODES else INACTIVE_SENTINEL))
            find_row(rytm["TAMLL"][BASE], TechId=cluster_id, MoId=mode).update(values(0))
            output = find_row(rytcm["OAR"][BASE], TechId=cluster_id, CommId=CROP_COMMODITIES[crop], MoId=mode)
            old_2021 = output["2021"]
            old_2025 = output["2025"]
            if mode in ACTIVE_MODES:
                for year in YEARS:
                    if int(year) >= 2021:
                        output[year] = output["2020"] * fofa_factor(crop, water, int(year))
            evidence.append({
                "cluster_id": cluster_id, "cluster": cluster_name, "mode": mode,
                "crop": crop, "water_regime": water, "inherited_input_label": level,
                "v30_status": "active_observed_cost_system" if mode in ACTIVE_MODES else "inactive_unsupported_pair",
                "activity_upper": 99999 if mode in ACTIVE_MODES else INACTIVE_SENTINEL,
                "yield_2020": output["2020"], "inherited_yield_2021": old_2021,
                "v30_yield_2021": output["2021"], "inherited_yield_2025": old_2025,
                "v30_yield_2025": output["2025"], "v30_yield_2050": output["2050"],
                "after_2020_rule": "FAO_BAU_2050_rebased_to_2020" if mode in ACTIVE_MODES else "inactive",
            })
    write_json(TARGET / "RYTM.json", rytm)
    write_json(TARGET / "RYTCM.json", rytcm)
    return evidence


def supersede_and_audit_history() -> None:
    path = TARGET / "data_sources" / "MODEL_MAP.csv"
    fields, rows = read_csv(path)
    supersessions = {
        "MAP_PHL_V24_CROP_OAR": "MAP_PHL_V30_CROP_YIELDS",
        "MAP_PHL_V29_CROP_COSTS": "MAP_PHL_V30_CROP_SYSTEMS",
        "MAP_PHL_V17_BUILT_UP_PATH": "MAP_PHL_V30_DYNAMIC_BUILT",
    }
    for item in rows:
        if item["map_id"] in supersessions:
            item["superseded_by"] = supersessions[item["map_id"]]
    write_csv(path, fields, rows)

    path = TARGET / "data_sources" / "CHANGES.csv"
    fields, rows = read_csv(path)
    for item in rows:
        if item["change_id"] == "CHG_PHL_V29_CROP_LAND_20260826":
            item["resolve_status"] = "implemented_preflight_and_cbc_complete"
            item["notes"] = "V29 source, pre-flight and solver run were completed and audited before v30 inheritance; v30 supersedes only the documented crop-system and built-land mappings."
    write_csv(path, fields, rows)


def install_provenance(population: dict[str, float], yield_rows: list[dict[str, Any]]) -> None:
    supersede_and_audit_history()
    source_fields, _ = read_csv(TARGET / "data_sources" / "SOURCES.csv")
    sources = [
        ["SRC_PHL_V30_PARENT", "This repository", "Philippines_v29 canonical source case", "Philippines_v29", "2020-2053", "Philippines", "complete immediate model lineage", "model source package", "WebAPP/DataStorage/Philippines_v29", "", "2026-08-26", "repository terms", "", "", "Complete v29 ledger and retained evidence copied before v30 additions; results and views are not inherited."],
        ["SRC_PHL_V30_PSA_PALAY_MICRODATA_CATALOG", "Philippine Statistics Authority", "Costs and Returns Survey of Palay Production", "2022 catalog", "2022", "Philippines", "parcel yield, irrigation, inputs, costs and survey weights", "farm microdata", "PSADA catalog 250", "https://psada.psa.gov.ph/catalog/250", "2026-08-26", "access agreement; redistribution restricted", "", "evidence/v30_yield_built/source_decision_register.csv", "Catalog inspected, but restricted microdata were not downloaded or used. It cannot support a redistributable paired high/low calibration in this package."],
        ["SRC_PHL_V30_PSA_TOMATO_MICRODATA_CATALOG", "Philippine Statistics Authority", "Costs and Returns Survey of Tomato Production", "2017 catalog", "2017", "six provinces", "farm yield, irrigation, inputs and costs", "farm microdata", "PSADA catalog 31", "https://psada.psa.gov.ph/catalog/31", "2026-08-26", "access agreement; redistribution restricted", "", "evidence/v30_yield_built/source_decision_register.csv", "Catalog inspected, but restricted microdata were not downloaded or used; it covers only tomato and cannot calibrate all modeled crops."],
        ["SRC_PHL_V30_FAO_FOFA2050_CROP", "Food and Agriculture Organization of the United Nations", "The future of food and agriculture - Alternative pathways to 2050 country crop data", "2018", "2012;2030;2035;2040;2050", "Philippines", "crop yield by crop, water regime and scenario", "tonnes/ha", "Country data by indicator: Crop production; CountryCode=PHL; Indicator=Crop yield; Scenario=Business As Usual", FOFA_SOURCE_URL, "2026-08-26", "FAO terms; required citation FAO 2018", FOFA_SOURCE_SHA256, "evidence/v30_yield_built/fao_fofa2050_phl_bau_crop_yield_extract.csv", "Selected Philippine BAU rows retained locally; annual model indices use linear interpolation between published milestones and extend the 2040-2050 slope through 2053."],
    ]
    append_ledger("SOURCES.csv", [row(item, source_fields) for item in sources], "source_id")

    fields, _ = read_csv(TARGET / "data_sources" / "ASSUMPTIONS.csv")
    assumptions = [
        ["ASM_PHL_V30_SINGLE_MANAGEMENT", "Only one cost-and-yield management system is operable per modeled crop and water regime until paired Philippine farm observations can support another.", "12 active systems", "classification", "SRC_PHL_V24_PSA_CROP_2020_2024;SRC_PHL_V29_PSA_VEGETABLE_COSTS;SRC_PHL_V29_SRA_SUGARCANE_COST;SRC_PHL_V30_PSA_PALAY_MICRODATA_CATALOG;SRC_PHL_V30_PSA_TOMATO_MICRODATA_CATALOG", "", "", "Prevents the optimizer choosing an unevidenced high/low combination whose GAEZ layers use different climate periods.", "The inherited high/full resource-cost mode is retained; twelve low-input modes use a 1e-6 operational-zero upper bound because sparse exact zeros are omitted by MUIO."],
        ["ASM_PHL_V30_YIELD_AFTER_2020", "Each active cluster/crop/water yield is initialized at the 2020 Philippine value and evolves with the corresponding FAO 2050 Business As Usual yield index.", "FAO BAU index rebased to 2020=1", "ratio", "SRC_PHL_V16_CROP_YIELD_INPUTS;SRC_PHL_V24_GAEZ_RELATIVE;SRC_PHL_V30_FAO_FOFA2050_CROP", "", "", "Uses an independent published projection rather than forcing or reproducing observed post-base outcomes.", "Milestones are interpolated linearly; the published 2040-2050 slope continues through 2053; maize and coconut use the published rainfed index for both model water regimes; the five-component other-crop aggregate is weighted by retained 2020 production."],
        ["ASM_PHL_V30_DYNAMIC_BUILT", "Built-up land equals population settlement land plus installed solar-PV and onshore-wind direct footprints in every solved year and scenario.", "PV 0.012; wind 0.010", "1000 km2/GW", "SRC_PSA_NAMRIA_LAND_ACCOUNTS_2020;SRC_PHL_V17_PSA_POP_SCENARIO2;SRC_PHL_V28_BOI_SOLAR_LAND;SRC_PHL_V28_NREL_WIND_LAND", "0", "national land endowment", "Makes renewable land compete during optimization through the existing built-up land class.", "The population component inherits the v17 path after subtracting the estimated 2020 PV/wind footprint so the 2020 total remains exactly observed."],
        ["ASM_PHL_V30_EXACT_BUILT_IDENTITY", "An equality UDC prevents surplus built-site production while the commodity balance makes population and renewable footprints mandatory.", "equality", "model relation", "SRC_PHL_V30_PARENT", "", "", "OSeMOSYS commodity balances alone are lower bounds; the equality makes the land account auditable and exact.", "No separate PV-site or wind-site technology is created."],
    ]
    append_ledger("ASSUMPTIONS.csv", [row(item, fields) for item in assumptions], "assumption_id")

    fields, _ = read_csv(TARGET / "data_sources" / "CALCULATIONS.csv")
    calculations = [
        ["CALC_PHL_V30_CROP_YIELDS", "active OAR[y]=OAR[2020]*FOFA_BAU_yield_index[y]/FOFA_BAU_yield_index[2020]; linear annual interpolation; y>2050 extends the 2040-2050 slope; unsupported paired mode TAMUL=1e-6", "SRC_PHL_V16_CROP_YIELD_INPUTS;SRC_PHL_V24_GAEZ_RELATIVE;SRC_PHL_V30_FAO_FOFA2050_CROP;SRC_PHL_V30_PSA_PALAY_MICRODATA_CATALOG;SRC_PHL_V30_PSA_TOMATO_MICRODATA_CATALOG", "ASM_PHL_V30_SINGLE_MANAGEMENT;ASM_PHL_V30_YIELD_AFTER_2020", "CALC_PHL_V24_CROP_OAR;CALC_PHL_V29_CROP_COSTS", "eight clusters; 24 inherited modes; 2020 OAR; FAO 2012/2030/2035/2040/2050 crop yields", "Mt/1000 km2; tonnes/ha; MUSD/1000 km2", "12 active crop/water systems with independent projected trajectories and 12 operationally inactive unsupported pairs", "model coefficients", "scripts/build_philippines_v30_yields_re_land.py", "v30", "Observed 2021-2024 yields are not installed; they remain available only as out-of-sample validation evidence."],
        ["CALC_PHL_V30_POPULATION_BUILT", "Bpop[y]=Bv29[y]*(10.2649-0.012*1.01207-0.010*0.4269)/10.2649", "SRC_PSA_NAMRIA_LAND_ACCOUNTS_2020;SRC_PHL_V17_PSA_POP_SCENARIO2;SRC_PHL_V30_PARENT", "ASM_PHL_V30_DYNAMIC_BUILT", "CALC_PHL_V17_BUILT_UP_PATH", "10.2649;1.01207;0.4269;0.012;0.010", "1000 km2;GW;1000 km2/GW", f"2020 population-built={POPULATION_BUILT_2020:.8f}", "1000 km2", "scripts/build_philippines_v30_yields_re_land.py", "v30", "Scaling preserves the inherited population path shape while avoiding double-counting the estimated existing RE footprint."],
        ["CALC_PHL_V30_RE_FOOTPRINT", "Btotal[y]=Bpop[y]+0.012*TotalCapacityPV[y]+0.010*TotalCapacityWind[y]", "SRC_PHL_V28_BOI_SOLAR_LAND;SRC_PHL_V28_NREL_WIND_LAND", "ASM_PHL_V30_DYNAMIC_BUILT;ASM_PHL_V30_EXACT_BUILT_IDENTITY", "CALC_PHL_V30_POPULATION_BUILT", "0.012;0.010", "1000 km2/GW", "scenario-endogenous annual built-land requirement", "1000 km2", "scripts/build_philippines_v30_yields_re_land.py", "v30", "Implemented both as built-site capacity demand and an exact equality UDC."],
    ]
    append_ledger("CALCULATIONS.csv", [row(item, fields) for item in calculations], "calculation_id")

    fields, _ = read_csv(TARGET / "data_sources" / "MODEL_MAP.csv")
    maps = [
        ["MAP_PHL_V30_CROP_SYSTEMS", "RYTM.json", "TAMLL/TAMUL; VC", "LNDAGRPHLC01-08", "1-24", "SC_0; policy rows inherit", "2020-2053", "modes 1,3,4,5,6,8,10,11,16,19,22,24 active; other crop modes TAMUL=1e-6", "1000 km2 and MUSD/1000 km2", "CALC_PHL_V30_CROP_YIELDS;ASM_PHL_V30_SINGLE_MANAGEMENT;CALC_PHL_V29_CROP_COSTS", "", "source;assumption;calculation", "Preserves crop and irrigation choice while removing only the unsupported input-level pair."],
        ["MAP_PHL_V30_CROP_YIELDS", "RYTCM.json", "OAR", "LNDAGRPHLC01-08 crop outputs", "active crop modes", "SC_0; policy rows inherit", "2020-2053", "2020 Philippine anchor with inherited spatial multiplier; 2021-2050 interpolated FAO BAU yield index; 2051-2053 extends the 2040-2050 slope", "Mt/1000 km2", "CALC_PHL_V30_CROP_YIELDS;ASM_PHL_V30_YIELD_AFTER_2020", "", "source;assumption;calculation", "Supersedes both the v24 observed 2021-2024 anchors and generic post-2024 productivity growth for active modes."],
        ["MAP_PHL_V30_DYNAMIC_BUILT", "genData.json;RYTM.json;RYTCM.json;RYTC.json;RYC.json;RYTCn.json;RYCn.json", "built-site OAR/AAD/ITCR; TAMLL/TAMUL; CAM/CCM/CNCM/UCC", "LNDBLTTOT;LBLTTOT;PHL_BUILT_SITE;LNDAGRPHLC01-08;PHL_POW_PP_SPV;PHL_POW_PP_WON;PHL_DYNAMIC_BUILT_LAND", "built mode 1; cluster mode 26", "all scenarios through inheritance", "2020-2053", "BuiltActivity=Bpop+0.012*PVTotalCapacity+0.010*WindTotalCapacity", "1000 km2", "CALC_PHL_V30_POPULATION_BUILT;CALC_PHL_V30_RE_FOOTPRINT;ASM_PHL_V30_EXACT_BUILT_IDENTITY", "", "source;assumption;calculation", "Uses existing built-up land and cluster allocation; no new land technology is created."],
        ["MAP_PHL_V30_CANONICAL_PACKAGE", "genData.json;root JSON;data_sources;documentation", "complete package identity and provenance", "Philippines_v30", "", "all", "2020-2053", "complete Philippines_v29 source/evidence plus v30 delta and regenerated standalone workbook", "package", "SRC_PHL_V30_PARENT;MAP_PHL_V30_CROP_SYSTEMS;MAP_PHL_V30_CROP_YIELDS;MAP_PHL_V30_DYNAMIC_BUILT", "", "model map", "The CSV ledgers remain authoritative; the XLSX is a formatted standalone view of all inherited and new records."],
    ]
    append_ledger("MODEL_MAP.csv", [row(item, fields) for item in maps], "map_id")

    fields, _ = read_csv(TARGET / "data_sources" / "CHANGES.csv")
    changes = [
        ["CHG_PHL_V30_YIELD_RE_LAND_20260826", "2026-08-26", "B", "Removed unsupported crop input-level operation, removed generic future yield growth, cleared the inherited policy-only idle-cropland cap, and made population/PV/wind footprints consume the existing physical built-up land account endogenously.", "genData.json;RYTM.json;RYTCM.json;RYTC.json;RYC.json;RYCTs.json;RYTCn.json;RYCn.json", "../documentation/MODEL_FIXES_YIELDS_RE_LAND_V30_2026-08-26.md;snapshots/philippines_v30_yield_built_inputs_2026-08-26.json;evidence/v30_yield_built", "MAP_PHL_V30_CROP_SYSTEMS;MAP_PHL_V30_CROP_YIELDS;MAP_PHL_V30_DYNAMIC_BUILT;MAP_PHL_V30_CANONICAL_PACKAGE", "source_implemented_preflight_pending", "Codex", "", "CBC must not run until the v30 feasibility and matrix gate passes; solver deadline is hard-limited to 300 seconds per scenario."],
        ["CHG_PHL_V30_LEDGER_AUDIT_20260826", "2026-08-26", "D", "Audited inherited v29 provenance, carried all six CSV ledgers and retained evidence forward, corrected the inherited v29 completion status, replaced the stale README, and generated a complete v30 workbook view.", "data_sources/SOURCES.csv;ASSUMPTIONS.csv;CALCULATIONS.csv;MODEL_MAP.csv;CHANGES.csv;GAPS.csv;RETAINED_EVIDENCE_MANIFEST.csv;PHILIPPINES_V30_CANONICAL_SCHEMA_LEDGER.xlsx", "data_sources/README.md", "MAP_PHL_V30_CANONICAL_PACKAGE", "complete", "Codex", "", "V29 had a complete authoritative CSV ledger but no v29-formatted workbook and retained a stale v18 README; v30 is standalone and does not require an earlier ledger."],
    ]
    append_ledger("CHANGES.csv", [row(item, fields) for item in changes], "change_id")

    fields, gaps = read_csv(TARGET / "data_sources" / "GAPS.csv")
    gaps.append({
        "item": "Paired Philippine crop yield and full-cost observations by management and water regime",
        "why_absent": "Available PSA farm microdata catalogs require a user agreement and restrict redistribution; accessible aggregates do not provide paired national systems for all modeled crops.",
        "upgrade_source": "Authorized PSA Costs and Returns microdata analysis producing redistributable weighted statistics, expanded across rice, maize, coconut, sugarcane, vegetables and other crops.",
        "priority": "high",
        "notes": "V30 keeps one evidenced full-cost system per crop/water regime and does not interpret mismatched GAEZ climate-period layers as a management response.",
    })
    write_csv(TARGET / "data_sources" / "GAPS.csv", fields, gaps)

    evidence_dir = TARGET / "data_sources" / "evidence" / "v30_yield_built"
    evidence_dir.mkdir(parents=True)
    decisions = [
        {"topic": "paired crop systems", "candidate": "PSA Palay 2022 microdata", "decision": "not used", "reason": "requires access agreement and redistribution is restricted", "locator": "https://psada.psa.gov.ph/catalog/250"},
        {"topic": "paired crop systems", "candidate": "PSA Tomato 2017 microdata", "decision": "not used", "reason": "restricted, six-province tomato-only sample cannot represent every modeled crop", "locator": "https://psada.psa.gov.ph/catalog/31"},
        {"topic": "crop yield initialization", "candidate": "inherited PSA 2020 anchor plus normalized GAEZ spatial pattern", "decision": "used", "reason": "source-traceable base-year initialization; no later observation is imposed", "locator": "snapshots/crop_yields_2020.json; snapshots/agriculture_v24_2026-08-25.json"},
        {"topic": "crop yield projection", "candidate": "FAO FOFA 2050 Philippines Business As Usual crop yields", "decision": "used", "reason": "independent published projection by crop and water regime through 2050", "locator": FOFA_SOURCE_URL},
        {"topic": "post-2020 yield", "candidate": "observed 2021-2024 anchors plus generic 1 percent growth", "decision": "rejected", "reason": "would use post-base outcomes and an unsourced generic trend", "locator": "ASM_PHL_V24_PRODUCTIVITY"},
        {"topic": "renewable footprint", "candidate": "v28 BOI PV and NREL direct wind factors", "decision": "used", "reason": "retained source evidence and agreed direct-footprint boundary", "locator": "SRC_PHL_V28_BOI_SOLAR_LAND;SRC_PHL_V28_NREL_WIND_LAND"},
    ]
    write_csv(evidence_dir / "source_decision_register.csv", list(decisions[0]), decisions)
    write_csv(evidence_dir / "yield_mode_mapping.csv", list(yield_rows[0]), yield_rows)
    fofa_rows = [
        {"country_code": "PHL", "scenario": "Business As Usual", "indicator": "Crop yield",
         "item": crop, "water_regime": water, "year": year, "tonnes_per_ha": value}
        for (crop, water), points in sorted(FOFA.items()) for year, value in sorted(points.items())
    ]
    write_csv(evidence_dir / "fao_fofa2050_phl_bau_crop_yield_extract.csv", list(fofa_rows[0]), fofa_rows)
    built_rows = [{
        "year": year, "inherited_total_path": population[year] * OBSERVED_BUILT_2020 / POPULATION_BUILT_2020,
        "population_component": population[year], "pv_footprint_1000km2_per_gw": PV_FOOTPRINT,
        "wind_footprint_1000km2_per_gw": WIND_FOOTPRINT,
    } for year in YEARS]
    write_csv(evidence_dir / "population_built_path.csv", list(built_rows[0]), built_rows)
    snapshot = {
        "schema": "philippines-v30-yield-built-inputs-v1", "date": "2026-08-26",
        "parent": "Philippines_v29", "active_crop_modes": sorted(ACTIVE_MODES),
        "inactive_crop_modes": sorted(INACTIVE_MODES), "inactive_upper_sentinel": INACTIVE_SENTINEL,
        "yield_rule": "2020 initialization; linear FAO BAU yield index 2021-2050; extend 2040-2050 slope through 2053",
        "fofa_source_url": FOFA_SOURCE_URL, "fofa_source_sha256": FOFA_SOURCE_SHA256,
        "fofa_selected_values": {f"{crop}:{water}": points for (crop, water), points in FOFA.items()},
        "built": {"observed_2020": OBSERVED_BUILT_2020, "pv_2020_gw": PV_2020, "wind_2020_gw": WIND_2020,
                  "pv_factor": PV_FOOTPRINT, "wind_factor": WIND_FOOTPRINT,
                  "population_2020": POPULATION_BUILT_2020, "population_path": population},
    }
    write_json(TARGET / "data_sources" / "snapshots" / "philippines_v30_yield_built_inputs_2026-08-26.json", snapshot)

    manifest_path = TARGET / "data_sources" / "evidence" / "RETAINED_EVIDENCE_MANIFEST.csv"
    mfields, manifest = read_csv(manifest_path)
    known = {item["relative_path"] for item in manifest}
    additions = list(evidence_dir.rglob("*")) + [TARGET / "data_sources" / "snapshots" / "philippines_v30_yield_built_inputs_2026-08-26.json"]
    for path in sorted(item for item in additions if item.is_file()):
        relative = path.relative_to(TARGET / "data_sources").as_posix()
        if relative not in known:
            manifest.append({"relative_path": relative, "size_bytes": str(path.stat().st_size), "sha256": sha256(path), "role": "Philippines v30 crop-yield and endogenous renewable-land evidence"})
    write_csv(manifest_path, mfields, manifest)


def write_documentation(population: dict[str, float]) -> None:
    text = f"""# Philippines v30 crop-yield and renewable-land implementation

Date: 2026-08-26  
Parent: `Philippines_v29`  
Status: source implementation; CBC is permitted only after the v30 pre-flight passes

## Crop systems and yields

V29's physical cropland and shared irrigation structure is retained. V30 operates one full-cost management system for each crop and water regime: modes `{','.join(map(str, sorted(ACTIVE_MODES)))}`. The twelve inherited low-input modes are operationally disabled with a `0.000001` thousand km2 upper bound. This is not a crop-share calibration: the optimizer remains free to choose crop, rainfed versus irrigated production, cluster, area and output. It removes only a high/low comparison that lacked paired Philippine yield-and-cost observations.

Only the active 2020 output ratios remain exactly the v24/v29 coefficients: the official Philippine production/area initialization normalized over the inherited GAEZ spatial pattern. From 2021 to 2050, each crop/water system follows the FAO *Future of Food and Agriculture 2050* Philippines Business As Usual yield index, rebased to its model 2020 value and linearly interpolated between FAO milestones. For 2051-2053 the published 2040-2050 annual slope continues. Thus observed 2021-2024 outcomes are not installed in the model; they remain out-of-sample checks. Restricted PSA palay and tomato microdata catalogs were inspected but not downloaded or redistributed; they are documented as future upgrade routes, not silently treated as evidence.

## Dynamic built-up land

No new land technology is created. The existing chain remains:

`PHL_LND -> LNDBLTTOT -> LBLTTOT -> LNDAGRPHLC01-08 mode 26 -> PHL_BUILT_SITE`.

Population, solar PV and onshore wind all consume `PHL_BUILT_SITE`. The exact annual equation, solved inside every scenario, is:

`BuiltActivity[y] = PopulationBuilt[y] + 0.012 * PVTotalCapacity[y] + 0.010 * WindTotalCapacity[y]`.

The factors are thousand km2/GW (1.2 ha/MW for PV and 1.0 ha/MW direct disturbance for wind). The population component follows the inherited v17 population-built path after removing the estimated 2020 renewable footprint. Its 2020 value is {POPULATION_BUILT_2020:.8f} thousand km2. Adding 2020 PV ({PV_2020} GW) and wind ({WIND_2020} GW) reconstructs the observed {OBSERVED_BUILT_2020} thousand km2 exactly.

Because the equation uses total installed capacity, RE requires land during optimization and automatically differs by scenario. Built expansion consumes the national `PHL_LND` endowment and therefore competes with cropland and other convertible classes. A same-year equality UDC prevents unused surplus built-site production.

V30 also clears an obsolete policy-scenario override inherited from v28 that capped idle cropland at 5.2877 thousand km2. BASE already left this choice open. All scenarios now inherit the same open BASE bound, so policy cases are not forced to cultivate surplus land.

## Provenance and run boundary

All six authoritative v29 CSV ledgers and every retained evidence file were copied before adding v30 records. The audit found that v29's CSV ledger was complete, but its README still described v18 and no v29-formatted workbook had been produced. V30 replaces the README, corrects the inherited v29 completion status, and generates `PHILIPPINES_V30_CANONICAL_SCHEMA_LEDGER.xlsx` as a formatted view of the complete standalone CSV ledger.

The v30 pre-flight checks identifier integrity, scenario inheritance, exact land equations, 2020 reconstruction, crop-mode activation, unchanged observed yield anchors, removal of future yield growth, analytical land headroom, complete provenance and generated matrices. CBC has a hard 300-second limit per scenario and is never called by the builder or pre-flight.
"""
    (TARGET / "documentation" / "MODEL_FIXES_YIELDS_RE_LAND_V30_2026-08-26.md").write_text(text, encoding="utf-8")
    readme = """# Philippines v30 canonical provenance ledger

This directory is the complete, standalone provenance record for `Philippines_v30`.

The authoritative tables are `SOURCES.csv`, `ASSUMPTIONS.csv`, `CALCULATIONS.csv`, `MODEL_MAP.csv`, `CHANGES.csv`, and `GAPS.csv`. They contain the full inherited history plus v30 changes; no earlier model package or ledger is required to interpret them. `PHILIPPINES_V30_CANONICAL_SCHEMA_LEDGER.xlsx` is a formatted review view generated from those six CSV files. Retained source material and hashes are listed in `evidence/RETAINED_EVIDENCE_MANIFEST.csv`.

V30-specific implementation documentation is in `../documentation/MODEL_FIXES_YIELDS_RE_LAND_V30_2026-08-26.md`; machine-readable inputs and decision records are in `snapshots/philippines_v30_yield_built_inputs_2026-08-26.json` and `evidence/v30_yield_built/`.
"""
    (TARGET / "data_sources" / "README.md").write_text(readme, encoding="utf-8")


def write_workbook() -> None:
    output = TARGET / "data_sources" / "PHILIPPINES_V30_CANONICAL_SCHEMA_LEDGER.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    for filename in LEDGERS:
        fields, rows = read_csv(TARGET / "data_sources" / filename)
        ws = wb.create_sheet(filename.removesuffix(".csv")[:31])
        ws.append(fields)
        for item in rows:
            ws.append([item[field] for field in fields])
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for index, field in enumerate(fields, 1):
            longest = max([len(field)] + [len(str(item[field])) for item in rows[:500]])
            ws.column_dimensions[ws.cell(1, index).column_letter].width = min(max(12, longest + 2), 55)
    wb.save(output)


def final_checks() -> None:
    gen = read_json(TARGET / "genData.json")
    if len(gen["osy-tech"]) != 160 or len(gen["osy-comm"]) != 73 or len(gen["osy-constraints"]) != 7:
        raise AssertionError("unexpected v30 set sizes")
    for filename in LEDGERS:
        fields, rows = read_csv(TARGET / "data_sources" / filename)
        key = fields[0]
        if len({item[key] for item in rows}) != len(rows):
            raise AssertionError(f"duplicate identifiers in {filename}")
    if not (TARGET / "data_sources" / "PHILIPPINES_V30_CANONICAL_SCHEMA_LEDGER.xlsx").is_file():
        raise AssertionError("missing v30 ledger workbook")


def main() -> int:
    copy_parent()
    population = population_built_path()
    install_general_data()
    install_built_land(population)
    yield_rows = install_yield_repair()
    install_provenance(population, yield_rows)
    write_documentation(population)
    write_workbook()
    final_checks()
    print(json.dumps({
        "status": "built", "source": str(SOURCE), "target": str(TARGET),
        "technologies": 160, "commodities": 73, "constraints": 7,
        "active_crop_modes": sorted(ACTIVE_MODES), "inactive_crop_modes": sorted(INACTIVE_MODES),
        "population_built_2020": POPULATION_BUILT_2020, "observed_built_2020": OBSERVED_BUILT_2020,
        "pv_land_factor": PV_FOOTPRINT, "wind_land_factor": WIND_FOOTPRINT,
        "cbc_invoked": False,
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
