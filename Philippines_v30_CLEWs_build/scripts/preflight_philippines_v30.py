#!/usr/bin/env python3
"""Fail-fast source, feasibility, provenance, and matrix gate for v30.

The gate generates all four scenario matrices but never invokes CBC.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
import sys
import types
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_philippines_v30_yields_re_land import (
    ACTIVE_MODES, BASE, BUILT, BUILT_SITE, CLUSTERS, CONSTRAINT, CROP_COMMODITIES,
    FOFA_SOURCE_SHA256, INACTIVE_MODES, INACTIVE_SENTINEL, MODES, OLD_OPTIONS,
    OTHER_LAND_TECH, POPULATION_BUILT_2020, PV, PV_FOOTPRINT, TARGET, WIND, WIND_FOOTPRINT, YEARS,
    fofa_factor,
)


ROOT = Path(__file__).resolve().parents[1]
STORAGE = ROOT / "WebAPP" / "DataStorage"
MODEL = ROOT / "WebAPP" / "SOLVERs" / "model.v.5.4.txt"
CASE = "Philippines_v30"
RUNS = {
    "SC_0": "BASE_V30", "SC_3hgjb": "COAL_PHASEOUT_V30",
    "SC_w03qj": "RE_V30", "SC_huc7i": "EV_V30",
}
LAND = "COM_fxuo5"
LAND_SUPPLY = "TEC_dgw1l"
PHYSICAL = {"TEC_phl_crop_v29", "TEC_mzni5", BUILT, "TEC_hjgww", "TEC_nozrj", "TEC_zty92", "TEC_6rzgn"}
PARAMETER_FILES = {
    "R.json", "RE.json", "RS.json", "RT.json", "RTSM.json", "RYC.json", "RYCTs.json",
    "RYCn.json", "RYDtb.json", "RYE.json", "RYS.json", "RYSeDt.json", "RYT.json",
    "RYTC.json", "RYTCM.json", "RYTCn.json", "RYTEM.json", "RYTM.json", "RYTTs.json", "RYTs.json",
}

dotenv_stub = types.ModuleType("dotenv")
dotenv_stub.load_dotenv = lambda *args, **kwargs: None
sys.modules.setdefault("dotenv", dotenv_stub)
sys.path.insert(0, str(ROOT / "API"))
from Classes.Case.DataFileClass import DataFile  # noqa: E402


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def find_row(rows: list[dict[str, Any]], **keys: Any) -> dict[str, Any]:
    found = [row for row in rows if all(row.get(key) == value for key, value in keys.items())]
    if len(found) != 1:
        raise AssertionError(f"expected one row for {keys}, found {len(found)}")
    return found[0]


def close(actual: float, expected: float, tolerance: float = 1e-8) -> None:
    if abs(actual - expected) > tolerance:
        raise AssertionError(f"{actual} != {expected} within {tolerance}")


def structure(case: Path) -> dict[str, Any]:
    gen = read_json(case / "genData.json")
    techs = {row["TechId"]: row for row in gen["osy-tech"]}
    comms = {row["CommId"]: row for row in gen["osy-comm"]}
    constraints = {row["ConId"]: row for row in gen["osy-constraints"]}
    if gen["osy-casename"] != CASE or len(techs) != 160 or len(comms) != 73 or len(constraints) != 7:
        raise AssertionError(f"wrong case identity or sizes: {gen['osy-casename']}, {len(techs)}, {len(comms)}, {len(constraints)}")
    if BUILT_SITE not in comms or CONSTRAINT not in constraints:
        raise AssertionError("v30 built-site objects are absent")
    if constraints[CONSTRAINT]["Tag"] != 1 or set(constraints[CONSTRAINT]["CM"]) != {BUILT, PV, WIND}:
        raise AssertionError("dynamic built constraint is not an exact three-technology identity")
    if BUILT_SITE not in techs[PV]["ITCR"] or BUILT_SITE not in techs[WIND]["ITCR"]:
        raise AssertionError("PV/wind do not declare installed-capacity land demand")
    if any(BUILT_SITE not in techs[cluster]["OAR"] for cluster in CLUSTERS):
        raise AssertionError("one or more land clusters cannot produce built-site service")

    errors = []
    scenarios = {row["ScenarioId"] for row in gen["osy-scenarios"]}
    for path in sorted(case.glob("*.json")):
        if path.name not in PARAMETER_FILES:
            continue
        payload = read_json(path)
        if not isinstance(payload, dict):
            continue
        for parameter, by_scenario in payload.items():
            if not isinstance(by_scenario, dict):
                continue
            if set(by_scenario) != scenarios:
                errors.append(f"{path.name}/{parameter} scenario keys differ")
            for scenario, rows in by_scenario.items():
                if not isinstance(rows, list):
                    continue
                for item in rows:
                    if item.get("TechId") and item["TechId"] not in techs:
                        errors.append(f"unknown technology {item['TechId']} in {path.name}")
                    if item.get("CommId") and item["CommId"] not in comms:
                        errors.append(f"unknown commodity {item['CommId']} in {path.name}")
                    if item.get("ConId") and item["ConId"] not in constraints:
                        errors.append(f"unknown constraint {item['ConId']} in {path.name}")
    if errors:
        raise AssertionError("; ".join(errors[:20]))
    return {"technologies": len(techs), "commodities": len(comms), "constraints": len(constraints), "references_valid": True}


def dynamic_built(case: Path) -> dict[str, Any]:
    rytm = read_json(case / "RYTM.json")
    rytcm = read_json(case / "RYTCM.json")
    rytc = read_json(case / "RYTC.json")
    ryc = read_json(case / "RYC.json")
    rytcn = read_json(case / "RYTCn.json")
    rycn = read_json(case / "RYCn.json")

    low = find_row(rytm["TAMLL"][BASE], TechId=BUILT, MoId=1)
    high = find_row(rytm["TAMUL"][BASE], TechId=BUILT, MoId=1)
    if any(float(low[y]) != 0 or float(high[y]) != 99999 for y in YEARS):
        raise AssertionError("obsolete fixed built-up path remains in activity bounds")

    for scenario in rytm["TAMUL"]:
        idle_low = find_row(rytm["TAMLL"][scenario], TechId=OTHER_LAND_TECH, MoId=2)
        idle_high = find_row(rytm["TAMUL"][scenario], TechId=OTHER_LAND_TECH, MoId=2)
        if scenario == BASE:
            if any(float(idle_low[y]) != 0 or float(idle_high[y]) != 99999 for y in YEARS):
                raise AssertionError("BASE idle-cropland choice is not open")
        elif any(idle_low[y] is not None or idle_high[y] is not None for y in YEARS):
            raise AssertionError(f"obsolete policy idle-cropland override remains in {scenario}")

    aad = find_row(ryc["AAD"][BASE], CommId=BUILT_SITE)
    ucc = find_row(rycn["UCC"][BASE], ConId=CONSTRAINT)
    for year in YEARS:
        close(float(aad[year]), float(ucc[year]))
    close(float(aad["2020"]), POPULATION_BUILT_2020)

    for tech, expected in ((PV, PV_FOOTPRINT), (WIND, WIND_FOOTPRINT)):
        itcr = find_row(rytc["ITCR"][BASE], TechId=tech, CommId=BUILT_SITE)
        ccm = find_row(rytcn["CCM"][BASE], TechId=tech, ConId=CONSTRAINT)
        for year in YEARS:
            close(float(itcr[year]), expected)
            close(float(ccm[year]), -expected)
    cam = find_row(rytcn["CAM"][BASE], TechId=BUILT, ConId=CONSTRAINT)
    if any(float(cam[y]) != 1 for y in YEARS):
        raise AssertionError("built activity coefficient is not +1")
    for parameter in ("CAM", "CCM", "CNCM"):
        for tech in (BUILT, PV, WIND):
            current = find_row(rytcn[parameter][BASE], TechId=tech, ConId=CONSTRAINT)
            expected = 1 if parameter == "CAM" and tech == BUILT else (-PV_FOOTPRINT if parameter == "CCM" and tech == PV else (-WIND_FOOTPRINT if parameter == "CCM" and tech == WIND else 0))
            if any(abs(float(current[y]) - expected) > 1e-12 for y in YEARS):
                raise AssertionError(f"wrong {parameter} coefficient for {tech}")

    service_rows = 0
    for cluster in CLUSTERS:
        for mode in MODES:
            output = find_row(rytcm["OAR"][BASE], TechId=cluster, CommId=BUILT_SITE, MoId=mode)
            expected = 1 if mode == 26 else 0
            if any(float(output[y]) != expected for y in YEARS):
                raise AssertionError(f"wrong built-site spatial link {cluster}/{mode}")
            service_rows += expected

    rc = read_json(case / "RYT.json")["RC"][BASE]
    pv_rc = find_row(rc, TechId=PV)
    wind_rc = find_row(rc, TechId=WIND)
    reconstructed = float(aad["2020"]) + PV_FOOTPRINT * float(pv_rc["2020"]) + WIND_FOOTPRINT * float(wind_rc["2020"])
    close(reconstructed, 10.2649, 1e-8)
    return {
        "population_2020": float(aad["2020"]), "pv_2020_gw": float(pv_rc["2020"]),
        "wind_2020_gw": float(wind_rc["2020"]), "reconstructed_built_2020": reconstructed,
        "pv_factor_1000km2_per_gw": PV_FOOTPRINT, "wind_factor_1000km2_per_gw": WIND_FOOTPRINT,
        "built_bounds_open": True, "policy_idle_cropland_bounds_inherit_base": True,
        "exact_equality": True, "spatial_service_links": service_rows,
    }


def crop_yields(case: Path) -> dict[str, Any]:
    parent = STORAGE / "Philippines_v29"
    rytm = read_json(case / "RYTM.json")
    ratios = read_json(case / "RYTCM.json")
    old_ratios = read_json(parent / "RYTCM.json")
    mapping = {mode: (crop, level, water) for mode, _ti, _tn, _ci, _cn, crop, level, water in OLD_OPTIONS}
    checked = 0
    for cluster in CLUSTERS:
        for mode in range(1, 25):
            upper = find_row(rytm["TAMUL"][BASE], TechId=cluster, MoId=mode)
            expected_upper = 99999 if mode in ACTIVE_MODES else INACTIVE_SENTINEL
            if any(abs(float(upper[y]) - expected_upper) > 1e-12 for y in YEARS):
                raise AssertionError(f"wrong crop-mode availability {cluster}/{mode}")
            crop, _level, water = mapping[mode]
            new = find_row(ratios["OAR"][BASE], TechId=cluster, CommId=CROP_COMMODITIES[crop], MoId=mode)
            old = find_row(old_ratios["OAR"][BASE], TechId=cluster, CommId=CROP_COMMODITIES[crop], MoId=mode)
            close(float(new["2020"]), float(old["2020"]), 1e-12)
            if mode in ACTIVE_MODES:
                for year in YEARS[1:]:
                    close(float(new[year]), float(new["2020"]) * fofa_factor(crop, water, int(year)), 1e-11)
                if abs(float(new["2021"]) - float(old["2021"])) < 1e-12 and abs(float(new["2024"]) - float(old["2024"])) < 1e-12:
                    raise AssertionError("post-base observed anchor series appears to remain")
            checked += 1
    return {
        "active_modes": sorted(ACTIVE_MODES), "inactive_modes": sorted(INACTIVE_MODES),
        "active_systems": len(ACTIVE_MODES), "inactive_unsupported_pairs": len(INACTIVE_MODES),
        "cluster_mode_rows_checked": checked, "only_2020_inherited_anchor": True,
        "projection": "FAO FOFA2050 Philippines Business As Usual; linear interpolation; 2040-2050 slope extended through 2053",
        "full_source_zip_sha256": FOFA_SOURCE_SHA256,
    }


def feasibility(case: Path) -> dict[str, Any]:
    """Necessary annual land-headroom test before any optimizer is allowed."""
    rytm = read_json(case / "RYTM.json")
    ryc = read_json(case / "RYC.json")
    ryt = read_json(case / "RYT.json")
    population = find_row(ryc["AAD"][BASE], CommId=BUILT_SITE)
    fixed_land = {"TEC_mzni5": 1, "TEC_hjgww": 1, "TEC_nozrj": 1, "TEC_zty92": 1, "TEC_6rzgn": 1, "TEC_phl_crop_v29": 1}
    floors = {tech: find_row(rytm["TAMLL"][BASE], TechId=tech, MoId=mode) for tech, mode in fixed_land.items()}
    pv_rc = find_row(ryt["RC"][BASE], TechId=PV)
    wind_rc = find_row(ryt["RC"][BASE], TechId=WIND)
    records = []
    minimum_headroom = float("inf")
    for year in YEARS:
        nonbuilt = sum(float(item[year]) for item in floors.values())
        mandatory_built = float(population[year]) + PV_FOOTPRINT * float(pv_rc[year]) + WIND_FOOTPRINT * float(wind_rc[year])
        required = nonbuilt + mandatory_built
        headroom = 295.8131 - required
        if headroom < -1e-8:
            raise AssertionError(f"necessary land floors exceed national endowment in {year}: {required}")
        minimum_headroom = min(minimum_headroom, headroom)
        records.append({"year": year, "nonbuilt_floor": nonbuilt, "mandatory_built_from_residual_capacity": mandatory_built, "required_floor": required, "headroom": headroom})
    close(records[0]["required_floor"], 295.8131, 1e-6)
    return {"necessary_condition": "pass", "national_endowment": 295.8131, "minimum_headroom": minimum_headroom, "annual": records}


def provenance(case: Path) -> dict[str, Any]:
    data = case / "data_sources"
    specs = {
        "SOURCES.csv": ("source_id", "SRC_PHL_V30_FAO_FOFA2050_CROP"),
        "ASSUMPTIONS.csv": ("assumption_id", "ASM_PHL_V30_YIELD_AFTER_2020"),
        "CALCULATIONS.csv": ("calculation_id", "CALC_PHL_V30_RE_FOOTPRINT"),
        "MODEL_MAP.csv": ("map_id", "MAP_PHL_V30_DYNAMIC_BUILT"),
        "CHANGES.csv": ("change_id", "CHG_PHL_V30_LEDGER_AUDIT_20260826"),
        "GAPS.csv": (None, None),
    }
    counts = {}
    for filename, (key, required) in specs.items():
        with (data / filename).open(newline="", encoding="utf-8") as handle:
            rows = list(csv.DictReader(handle))
        if key:
            ids = [row[key] for row in rows]
            if len(ids) != len(set(ids)) or required not in ids:
                raise AssertionError(f"bad or incomplete {filename}")
        counts[filename] = len(rows)
    workbook = data / "PHILIPPINES_V30_CANONICAL_SCHEMA_LEDGER.xlsx"
    wb = load_workbook(workbook, read_only=True)
    if set(wb.sheetnames) != {name.removesuffix(".csv") for name in specs}:
        raise AssertionError("ledger workbook sheets do not mirror all six CSV ledgers")
    manifest_path = data / "evidence" / "RETAINED_EVIDENCE_MANIFEST.csv"
    with manifest_path.open(newline="", encoding="utf-8") as handle:
        manifest = list(csv.DictReader(handle))
    verified = 0
    for item in manifest:
        if not item["relative_path"].startswith("evidence/v30_yield_built/"):
            continue
        path = data / item["relative_path"]
        if not path.is_file() or sha256(path) != item["sha256"] or path.stat().st_size != int(item["size_bytes"]):
            raise AssertionError(f"evidence manifest mismatch: {item['relative_path']}")
        verified += 1
    if verified < 4:
        raise AssertionError("v30 retained evidence is incomplete")
    readme = (data / "README.md").read_text(encoding="utf-8")
    if "v30" not in readme.lower() or "standalone" not in readme.lower():
        raise AssertionError("stale provenance README")
    return {"ledger_rows": counts, "workbook_sheets": wb.sheetnames, "v30_evidence_files_verified": verified, "standalone": True}


def ensure_run(model: DataFile, run: str, active: str) -> None:
    path = STORAGE / model.case / "res" / run
    scenarios = [{
        "ScenarioId": item["ScenarioId"], "Scenario": item["Scenario"], "Desc": item.get("Desc", ""),
        "Active": item["ScenarioId"] == BASE or item["ScenarioId"] == active,
    } for item in model.genData["osy-scenarios"]]
    if path.is_dir():
        found = False
        for item in model.resData["osy-cases"]:
            if item["Case"] == run:
                item["Scenarios"] = scenarios
                found = True
                break
        if not found:
            raise AssertionError(f"run directory exists without a run registry entry: {run}")
        write_json(STORAGE / model.case / "view" / "resData.json", model.resData)
        return
    response = model.createCaseRun(run, {
        "Case": run, "CaseId": f"RUN_{run}", "Desc": "Philippines v30 pre-solver feasibility matrix",
        "Runtime": date.today().isoformat(), "Scenarios": scenarios,
    })
    if response.get("status_code") != "success":
        raise RuntimeError(json.dumps(response, indent=2))


def matrices(case_name: str) -> dict[str, Any]:
    model = DataFile(case_name)
    results = {}
    for scenario, run in RUNS.items():
        ensure_run(model, run, scenario)
        path = STORAGE / case_name / "res" / run
        model.generateDatafile(run)
        model.preprocessData(path / "data.txt", path / "data_processed.txt")
        checked = subprocess.run(
            ["glpsol", "--check", "-m", str(MODEL), "-d", str(path / "data_processed.txt"), "--wlp", str(path / "lp.lp")],
            capture_output=True, text=True,
        )
        output = checked.stdout + checked.stderr
        (path / "glpsol_check.log").write_text(output, encoding="utf-8")
        if checked.returncode != 0 or "Model has been successfully generated" not in output:
            raise AssertionError(f"GLPK matrix failure for {run}:\n{output[-5000:]}")
        dimensions = {}
        for label, pattern in (("rows", r"Number of rows\s*=\s*([\d,]+)"), ("columns", r"Number of columns\s*=\s*([\d,]+)"), ("nonzeros", r"Number of non-zeros \(matrix\)\s*=\s*([\d,]+)")):
            match = re.search(pattern, output)
            dimensions[label] = int(match.group(1).replace(",", "")) if match else None
        if dimensions["rows"] is None or dimensions["rows"] < 460000 or dimensions["columns"] is None or dimensions["columns"] < 500000 or dimensions["nonzeros"] is None or dimensions["nonzeros"] < 8000000:
            raise AssertionError(f"{run} lost the BASE parameter layer: undersized matrix {dimensions}")
        forbidden = [name for name in ("cbc.log", "results.txt", "optimization_record.json") if (path / name).exists()]
        if forbidden:
            raise AssertionError(f"optimizer artifacts predate approval in {run}: {forbidden}")
        results[run] = {"scenario": scenario, "matrix": dimensions, "lp_sha256": sha256(path / "lp.lp"), "cbc_invoked": False}
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--case", default=CASE)
    parser.add_argument("--skip-matrices", action="store_true")
    args = parser.parse_args()
    case = STORAGE / args.case
    report_path = case / "documentation" / "preflight_v30.json"
    report: dict[str, Any] = {"case": args.case, "status": "running", "optimizer_runs": 0, "cbc_invoked": False, "checks": {}}
    try:
        report["checks"]["structure"] = structure(case)
        report["checks"]["dynamic_built_land"] = dynamic_built(case)
        report["checks"]["crop_yields"] = crop_yields(case)
        report["checks"]["land_feasibility"] = feasibility(case)
        report["checks"]["provenance"] = provenance(case)
        if not args.skip_matrices:
            report["checks"]["scenario_matrices"] = matrices(args.case)
        report["status"] = "pass"
        report["interpretation"] = "All v30 source and necessary-feasibility checks passed. Four scenario matrices were generated without invoking CBC." if not args.skip_matrices else "Static checks passed; matrices were explicitly skipped."
        write_json(report_path, report)
        print(json.dumps(report, indent=2))
        return 0
    except Exception as error:
        report["status"] = "fail"
        report["error"] = str(error)
        write_json(report_path, report)
        print(json.dumps(report, indent=2), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
